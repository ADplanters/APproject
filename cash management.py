import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 1. 데이터 구조 정의
data = {
    "수익_대분류": ["주수입", "주수입", "부수입", "부수입", "기타수입", "기타수입", "", ""],
    "수익_항목": ["본업 급여", "사업 매출", "알바/과외", "블로그/앱테크", "상여/성과급", "중고거래/기타", "", ""],
    "수익_목표": [0, 0, 0, 0, 0, 0, "", ""],
    "수익_실제": [0, 0, 0, 0, 0, 0, "", ""],
    "공백1": [""] * 8,
    "고정_대분류": ["필수 고정비", "필수 고정비", "필수 고정비", "필수 고정비", "필수 고정비", "계획저축", "계획저축", "계획저축"],
    "고정_항목": ["주거비(월세)", "공과금/관리비", "통신비/인터넷", "보험료", "구독 서비스", "저축(적금)", "투자(주식)", "대출 원리금"],
    "고정_예산": [0, 0, 0, 0, 0, 0, 0, 0],
    "고정_실제": [0, 0, 0, 0, 0, 0, 0, 0],
    "공백2": [""] * 8,
    "변동_대분류": ["필수 변동비", "필수 변동비", "필수 변동비", "필수 변동비", "자율지출", "자율지출", "자율지출", "자율지출"],
    "변동_항목": ["식비(장보기)", "생필품 구매", "교통비/주유", "의료비/약값", "외식/카페", "문화/여가", "품위유지/옷", "경조사/선물"],
    "변동_예산": [0, 0, 0, 0, 0, 0, 0, 0],
    "변동_실제": [0, 0, 0, 0, 0, 0, 0, 0]
}

df = pd.DataFrame(data)

# 2. 엑셀 파일 생성 및 스타일링 시작
with pd.ExcelWriter("자금_흐름_대시보드.xlsx", engine="openpyxl") as writer:
    df.to_excel(writer, startrow=4, index=False)
    worksheet = writer.sheets['Sheet1']
    
    # --- 디자인 서식 정의 ---
    title_font = Font(name='맑은 고딕', size=16, bold=True, color='1F4E78')
    header_font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    data_font = Font(name='맑은 고딕', size=10, color='333333')
    
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    thin_side = Side(style='thin', color='E0E0E0')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # --- 1. 상단 타이틀 추가 ---
    worksheet['A2'] = "📊 개인 자금 흐름 대시보드"
    worksheet['A2'].font = title_font
    worksheet.row_dimensions[2].height = 25
    
    # --- 2. 기본 데이터 스타일링 ---
    worksheet.row_dimensions[5].height = 24
    for col_num in range(1, len(df.columns) + 1):
        col_letter = get_column_letter(col_num)
        if col_num in [5, 10]:
            worksheet.column_dimensions[col_letter].width = 3
            for r in range(5, 6 + len(df)):
                worksheet.cell(row=r, column=col_num, value="").fill = white_fill
            continue
            
        header_cell = worksheet.cell(row=5, column=col_num)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = center_align
        header_cell.border = thin_border
        
        for row_num in range(6, 6 + len(df)):
            data_cell = worksheet.cell(row=row_num, column=col_num)
            data_cell.font = data_font
            data_cell.border = thin_border
            worksheet.row_dimensions[row_num].height = 20
            
            if isinstance(data_cell.value, (int, float)):
                data_cell.number_format = '#,##0'
                data_cell.alignment = right_align
            elif "대분류" in df.columns[col_num-1]:
                data_cell.alignment = center_align
            else:
                data_cell.alignment = left_align

    # --- 3. 각 섹션별 SUM 합계 행 추가 (14행) ---
    total_row = 14
    worksheet.row_dimensions[total_row].height = 24
    total_font = Font(name='맑은 고딕', size=10, bold=True, color='111111')
    total_fill = PatternFill(start_color='EAEDED', end_color='EAEDED', fill_type='solid')
    total_border = Border(top=Side(style='thin', color='B0B0B0'), bottom=Side(style='double', color='2C3E50'),
                          left=Side(style='thin', color='E0E0E0'), right=Side(style='thin', color='E0E0E0'))
    
    worksheet.cell(row=total_row, column=2, value="수익 합계").alignment = center_align
    worksheet.cell(row=total_row, column=3, value="=SUM(C6:C13)").number_format = '#,##0'
    worksheet.cell(row=total_row, column=4, value="=SUM(D6:D13)").number_format = '#,##0'
    
    worksheet.cell(row=total_row, column=7, value="고정비 합계").alignment = center_align
    worksheet.cell(row=total_row, column=8, value="=SUM(H6:H13)").number_format = '#,##0'
    worksheet.cell(row=total_row, column=9, value="=SUM(I6:I13)").number_format = '#,##0'
    
    worksheet.cell(row=total_row, column=12, value="변동비 합계").alignment = center_align
    worksheet.cell(row=total_row, column=13, value="=SUM(M6:M13)").number_format = '#,##0'
    worksheet.cell(row=total_row, column=14, value="=SUM(N6:N13)").number_format = '#,##0'
    
    for col_num in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=total_row, column=col_num)
        cell.font = total_font
        cell.border = total_border
        if col_num in [5, 10]:
            cell.value = ""
            cell.fill = white_fill
        else:
            cell.fill = total_fill
            if col_num in [3, 4, 8, 9, 13, 14]:
                cell.alignment = right_align

    # --- 4. 실생활용 '최종 요약 통계 카드' 추가 (16~17행) ---
    worksheet.row_dimensions[16].height = 20
    worksheet.row_dimensions[17].height = 28
    card_title_font = Font(name='맑은 고딕', size=10, color='555555')
    card_val_font = Font(name='맑은 고딕', size=13, bold=True, color='111111')
    
    fill_income = PatternFill(start_color='E8F8F5', end_color='E8F8F5', fill_type='solid')
    fill_expense = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    fill_net = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
    box_border = Border(left=Side(style='thin', color='B0B0B0'), right=Side(style='thin', color='B0B0B0'),
                        top=Side(style='thin', color='B0B0B0'), bottom=Side(style='thin', color='B0B0B0'))
    
    worksheet.merge_cells('B16:C16'); worksheet['B16'] = "💰 이번 달 총 수입 (실제)"; worksheet['B16'].font = card_title_font; worksheet['B16'].alignment = center_align
    worksheet.merge_cells('B17:C17'); worksheet['B17'] = "=D14"; worksheet['B17'].font = card_val_font; worksheet['B17'].alignment = center_align; worksheet['B17'].number_format = '#,##0'
    
    worksheet.merge_cells('G16:H16'); worksheet['G16'] = "💸 이번 달 총 지출 (고정+변동)"; worksheet['G16'].font = card_title_font; worksheet['G16'].alignment = center_align
    worksheet.merge_cells('G17:H17'); worksheet['G17'] = "=I14+N14"; worksheet['G17'].font = card_val_font; worksheet['G17'].alignment = center_align; worksheet['G17'].number_format = '#,##0'
    
    worksheet.merge_cells('L16:M16'); worksheet['L16'] = "🪙 이번 달 최종 남은 잔액"; worksheet['L16'].font = Font(name='맑은 고딕', size=10, bold=True, color='2E4053'); worksheet['L16'].alignment = center_align
    worksheet.merge_cells('L17:M17'); worksheet['L17'] = "=B17-G17"; worksheet['L17'].font = Font(name='맑은 고딕', size=14, bold=True, color='1F4E78'); worksheet['L17'].alignment = center_align; worksheet['L17'].number_format = '#,##0'
    
    for r in [16, 17]:
        for c in [2, 3]: worksheet.cell(row=r, column=c).fill = fill_income; worksheet.cell(row=r, column=c).border = box_border
        for c in [7, 8]: worksheet.cell(row=r, column=c).fill = fill_expense; worksheet.cell(row=r, column=c).border = box_border
        for c in [12, 13]: worksheet.cell(row=r, column=c).fill = fill_net; worksheet.cell(row=r, column=c).border = box_border

    # --- 💡 5. 핵심 업그레이드: '마인드셋 리마인드 & 비전 보드' 영역 추가 (19~25행) ---
    # 대구분 타이틀 (19행)
    worksheet.merge_cells('B19:N19')
    worksheet['B19'] = "📝 마인드셋 & 피드백 (스스로 돌아보며 자산 키우기)"
    worksheet['B19'].font = Font(name='맑은 고딕', size=12, bold=True, color='1F4E78')
    worksheet['B19'].alignment = Alignment(horizontal='left', vertical='center')
    worksheet.row_dimensions[19].height = 24
    
    # 카드별 서식 지정
    feedback_title_font = Font(name='맑은 고딕', size=11, bold=True, color='2C3E50')
    placeholder_font = Font(name='맑은 고딕', size=10, italic=True, color='7F8C8D')
    content_align = Alignment(horizontal='left', vertical='top', wrap_text=True) # 왼쪽 상단 정렬 및 줄바꿈 활성화
    
    # [1번 카드] 이번 달 잘한 점 (B~D열 / 21~25행)
    worksheet.merge_cells('B21:D21'); worksheet['B21'] = "🎉 이번 달 칭찬해요! (잘한 점)"; worksheet['B21'].font = feedback_title_font; worksheet['B21'].alignment = center_align
    worksheet.merge_cells('B22:D25'); worksheet['B22'] = "- 예산 범위 내에서 지출 방어 성공!\n- 무지출 데이 4회 달성\n- 충동구매 욕구를 잘 참아냄"
    worksheet['B22'].font = placeholder_font; worksheet['B22'].alignment = content_align
    fill_well = PatternFill(start_color='E8F8F5', end_color='E8F8F5', fill_type='solid') # 민트 계열
    for r in range(21, 26):
        for c in range(2, 5):
            cell = worksheet.cell(row=r, column=c); cell.border = box_border
            if r == 21: cell.fill = fill_well

    # [2번 카드] 다음 달 목표 및 개선점 (G~I열 / 21~25행)
    worksheet.merge_cells('G21:I21'); worksheet['G21'] = "🎯 다음 달 개선점 & 핵심 목표"; worksheet['G21'].font = feedback_title_font; worksheet['G21'].alignment = center_align
    worksheet.merge_cells('G22:I25'); worksheet['G22'] = "- 배달 음식을 너무 자주 먹음 (주 3회 -> 주 1회 목표)\n- 고정비 다이어트 필요 (알뜰폰 요금제 비교 후 변경)\n- 변동비 예산 10% 감축 도전"
    worksheet['G22'].font = placeholder_font; worksheet['G22'].alignment = content_align
    fill_improve = PatternFill(start_color='FEF9E7', end_color='FEF9E7', fill_type='solid') # 옐로우 계열
    for r in range(21, 26):
        for c in range(7, 10):
            cell = worksheet.cell(row=r, column=c); cell.border = box_border
            if r == 21: cell.fill = fill_improve

    # [3번 카드] 나의 최종 목표 (L~N열 / 21~25행)
    worksheet.merge_cells('L21:N21'); worksheet['L21'] = "🏆 나의 최종 자산 목표 (Vision)"; worksheet['L21'].font = feedback_title_font; worksheet['L21'].alignment = center_align
    worksheet.merge_cells('L22:N25'); worksheet['L22'] = "- 3년 내 투자 시드머니 5,000만 원 달성하기\n- 매달 소득의 50% 이상 무조건 저축/투자하기\n- 경제적 자유를 얻어 건강하고 여유로운 삶 구축"
    worksheet['L22'].font = placeholder_font; worksheet['L22'].alignment = content_align
    fill_vision = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid') # 블루 계열
    for r in range(21, 26):
        for c in range(12, 15):
            cell = worksheet.cell(row=r, column=c); cell.border = box_border
            if r == 21: cell.fill = fill_vision

    # 행 높이 설정
    worksheet.row_dimensions[21].height = 22
    for r in range(22, 26): worksheet.row_dimensions[r].height = 20

    # --- 6. 열 너비 맞춤 ---
    for col_num in range(1, len(df.columns) + 1):
        if col_num in [5, 10]: continue
        col_letter = get_column_letter(col_num)
        max_len = 0
        for row_num in range(5, 15):
            val = worksheet.cell(row=row_num, column=col_num).value
            if val:
                length = sum(2 if ord(char) > 128 else 1 for char in str(val))
                if length > max_len: max_len = length
        worksheet.column_dimensions[col_letter].width = max(max_len + 6, 14)

print("엑셀 파일이 성공적으로 생성되었습니다!")